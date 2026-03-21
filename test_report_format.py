"""
Local formatting test — generates sample_report.html so you can preview
the new HTML report design before it goes to production.

Run:
    ~/Downloads/attendance_env/bin/python test_report_format.py
"""

import base64
import io
import os
import subprocess
from datetime import date, datetime
from html import escape

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Embed logo as base64 (self-contained HTML) ────────────────────────────────

def _logo_base64():
    """Return a base64 data-URI for the TechSur logo, or None if not found."""
    logo_path = os.path.join(os.path.dirname(__file__), "techsur_logo.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            return "data:image/png;base64," + base64.b64encode(f.read()).decode()
    return None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _pct_badge(val):
    try:
        v = float(val)
        if v == 0:    cls, label = "badge-red",    f"{v:.0f}%"
        elif v < 80:  cls, label = "badge-orange",  f"{v:.0f}%"
        elif v < 100: cls, label = "badge-yellow",  f"{v:.0f}%"
        else:         cls, label = "badge-green",   f"{v:.0f}%"
        return f'<span class="badge {cls}">{label}</span>'
    except (TypeError, ValueError):
        return escape(str(val))


def _build_table(df, show_manager=True):
    has_mgr = show_manager and "Manager" in df.columns
    mgr_th  = "<th style='text-align:left'>Manager</th>" if has_mgr else ""

    rows = ""
    for i, (_, row) in enumerate(df.iterrows(), 1):
        mgr_cell = (
            f'<td class="muted">{escape(str(row.get("Manager", "")))}</td>'
            if has_mgr else ""
        )
        pct = row["Attendance %"]
        rows += (
            f"<tr>"
            f'<td class="num light">{i}</td>'
            f'<td class="emp">{escape(str(row["_name"]))}</td>'
            f"{mgr_cell}"
            f'<td class="num">{int(row["Days Present"])}</td>'
            f'<td class="num">{int(row["Days Absent"])}</td>'
            f'<td class="num light">{int(row["Total Weekdays"])}</td>'
            f'<td class="center">{_pct_badge(pct)}</td>'
            f"</tr>"
        )

    return (
        f'<table><thead><tr>'
        f'<th style="width:36px">#</th>'
        f'<th style="text-align:left">Employee</th>'
        f"{mgr_th}"
        f'<th>Present</th><th>Absent</th><th>Total</th><th>Attendance</th>'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


# ── Main HTML generator ────────────────────────────────────────────────────────

def generate_report_html(unique_days, zero_df, start, end, total_weekdays):
    period       = f"{start.strftime('%B %d')} \u2013 {end.strftime('%B %d, %Y')}"
    generated_on = datetime.now().strftime("%B %d, %Y")

    total_emp  = len(unique_days)
    avg_pct    = unique_days["Attendance %"].mean() if total_emp else 0.0
    at_risk    = int((unique_days["Attendance %"] < 80).sum())
    zero_count = len(zero_df)

    # ── Logo ───────────────────────────────────────────────────────────────
    logo_src = _logo_base64()
    if logo_src:
        logo_html = f'<img src="{logo_src}" alt="TechSur" style="height:52px;display:block;">'
    else:
        # CSS fallback that matches the brand look
        logo_html = (
            '<div style="display:flex;flex-direction:column;align-items:flex-start;">'
            '<div style="font-size:28px;font-weight:900;color:#F0B429;letter-spacing:2px;'
            'font-family:\'Segoe UI\',Arial,sans-serif;line-height:1;">TECHSUR</div>'
            '<div style="font-size:8px;color:rgba(255,255,255,0.55);letter-spacing:2.5px;'
            'font-weight:600;text-transform:uppercase;margin-top:3px;">PASSION MEETS TECHNOLOGY</div>'
            '</div>'
        )

    # ── All Employees table ────────────────────────────────────────────────
    sort_cols  = ["Manager", "Attendance %"] if "Manager" in unique_days.columns else ["Attendance %"]
    all_sorted = unique_days.sort_values(sort_cols, ascending=True)
    all_table  = _build_table(all_sorted, show_manager=True)

    # ── By-Manager sections (collapsible) ─────────────────────────────────
    mgr_sections = ""
    if "Manager" in unique_days.columns:
        named = sorted([
            m for m in unique_days["Manager"].dropna().unique()
            if m not in ("No Manager", "Unknown / Not Mapped", "")
        ])
        for mgr in named:
            team     = unique_days[unique_days["Manager"] == mgr].copy()
            avg      = team["Attendance %"].mean()
            n        = len(team)
            risk_n   = int((team["Attendance %"] < 80).sum())
            tbl      = _build_table(team.sort_values("Attendance %"), show_manager=False)
            risk_pill = (
                f'<span class="risk-pill">{risk_n} at risk</span>'
                if risk_n > 0 else ""
            )
            mgr_sections += (
                f'<details class="mgr-block">'
                f'<summary class="mgr-bar">'
                f'<span class="mgr-left">'
                f'<span class="chevron">&#9656;</span>'
                f'<span class="mgr-name-title">{escape(mgr)}</span>'
                f'</span>'
                f'<span class="mgr-meta">'
                f'{n} employee{"s" if n != 1 else ""}'
                f'&nbsp;&nbsp;&bull;&nbsp;&nbsp;Avg: {_pct_badge(avg)}'
                f'&nbsp;&nbsp;{risk_pill}'
                f'</span>'
                f'</summary>'
                f'<div class="mgr-table-wrap">{tbl}</div>'
                f'</details>'
            )

        no_mgr = unique_days[
            unique_days["Manager"].isin(["No Manager", "Unknown / Not Mapped"])
        ].copy()
        if not no_mgr.empty:
            tbl = _build_table(no_mgr.sort_values("Attendance %"), show_manager=False)
            mgr_sections += (
                f'<details class="mgr-block">'
                f'<summary class="mgr-bar" style="border-left-color:#555;">'
                f'<span class="mgr-left">'
                f'<span class="chevron">&#9656;</span>'
                f'<span class="mgr-name-title" style="color:#aaa;">No Manager Assigned</span>'
                f'</span>'
                f'<span class="mgr-meta">{len(no_mgr)} employee{"s" if len(no_mgr) != 1 else ""}</span>'
                f'</summary>'
                f'<div class="mgr-table-wrap">{tbl}</div>'
                f'</details>'
            )

    # ── 0 Attendance section ───────────────────────────────────────────────
    zero_section = ""
    if not zero_df.empty:
        zero_tbl = _build_table(zero_df.sort_values("_name"), show_manager=True)
        zero_section = (
            f'<div class="section zero-section">'
            f'<div class="section-header">'
            f'<h2 style="color:#C0392B;">&#9888; Zero Attendance</h2>'
            f'<span class="pill" style="background:#FDEDEC;color:#C0392B;">{len(zero_df)}</span>'
            f'</div>'
            f'<p class="note">These employees have a DataWatch badge assigned but no recorded'
            f' office entries for this period.</p>'
            f'{zero_tbl}</div>'
        )

    # ── Stat cards ─────────────────────────────────────────────────────────
    avg_color  = "#27AE60" if avg_pct >= 80 else "#E67E22" if avg_pct >= 60 else "#E74C3C"
    risk_color = "#E74C3C" if at_risk    > 0 else "#27AE60"
    zero_color = "#E74C3C" if zero_count > 0 else "#27AE60"

    cards = (
        f'<div class="card">'
        f'<div class="stat">{total_emp}</div>'
        f'<div class="stat-label">Employees Tracked</div></div>'

        f'<div class="card">'
        f'<div class="stat" style="color:{avg_color}">{avg_pct:.1f}%</div>'
        f'<div class="stat-label">Avg Attendance</div></div>'

        f'<div class="card">'
        f'<div class="stat" style="color:{risk_color}">{at_risk}</div>'
        f'<div class="stat-label">At Risk (&lt;80%)</div></div>'

        f'<div class="card">'
        f'<div class="stat" style="color:{zero_color}">{zero_count}</div>'
        f'<div class="stat-label">Zero Attendance</div></div>'
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TechSur Attendance Report &ndash; {period}</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    background: #F0F2F5;
    color: #2D2D2D;
    padding: 28px 16px;
    font-size: 14px;
    line-height: 1.5;
  }}

  .report {{
    max-width: 980px;
    margin: 0 auto;
    background: #fff;
    border-radius: 10px;
    box-shadow: 0 2px 16px rgba(0,0,0,0.09);
    overflow: hidden;
  }}

  /* ── Header (dark — logo lives here, keep brand) ─────────── */
  .header {{
    background: #1A1A1C;
    padding: 24px 40px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 3px solid #F0B429;
  }}
  .header-right {{ text-align: right; }}
  .header-right .report-title {{
    font-size: 19px; font-weight: 700; color: #FFFFFF;
    letter-spacing: 0.3px; margin-bottom: 4px;
  }}
  .header-right .report-meta {{ font-size: 12px; color: rgba(255,255,255,0.5); }}
  .header-right .report-meta strong {{ color: #F0B429; font-weight: 600; }}

  /* ── Stat cards ──────────────────────────────────────────── */
  .cards {{
    display: flex;
    background: #FAFAF8;
    border-bottom: 1px solid #ECECEC;
  }}
  .card {{
    flex: 1; padding: 20px 16px; text-align: center;
    border-right: 1px solid #ECECEC;
  }}
  .card:last-child {{ border-right: none; }}
  .stat       {{ font-size: 32px; font-weight: 700; color: #2D2D2D; line-height: 1; margin-bottom: 5px; }}
  .stat-label {{ font-size: 10px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.8px; color: #9E9E9E; }}

  /* ── Section ─────────────────────────────────────────────── */
  .section {{
    padding: 26px 40px 34px;
    border-bottom: 1px solid #ECECEC;
  }}
  .section:last-child {{ border-bottom: none; }}
  .section-header {{
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 16px;
  }}
  .section-header h2 {{
    font-size: 15px; font-weight: 700; color: #2D2D2D;
    display: flex; align-items: center; gap: 10px;
  }}
  .section-header h2::before {{
    content: ''; display: inline-block;
    width: 4px; height: 18px; border-radius: 2px;
    background: #F0B429;
  }}
  .pill {{
    background: #FFF3CD; color: #856404;
    font-size: 11px; font-weight: 700;
    padding: 3px 10px; border-radius: 12px;
  }}
  .note {{ font-size: 12px; color: #9E9E9E; margin-bottom: 12px; }}

  /* ── Table ───────────────────────────────────────────────── */
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  thead th {{
    background: #3D3A35; color: #FFFFFF;
    padding: 10px 14px;
    font-size: 11px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.6px;
    text-align: center; white-space: nowrap;
  }}
  thead th[style*="left"] {{ text-align: left; }}
  tbody tr:nth-child(even) td {{ background: #FAFAF8; }}
  tbody tr:hover td {{ background: #FFF8E8 !important; }}
  tbody td {{
    padding: 9px 14px;
    border-bottom: 1px solid #F0EFED;
    vertical-align: middle;
  }}
  td.num    {{ text-align: center; font-variant-numeric: tabular-nums; color: #4A4A4A; }}
  td.center {{ text-align: center; }}
  td.emp    {{ font-weight: 600; color: #2D2D2D; }}
  td.muted  {{ color: #888; font-size: 12px; }}
  td.light  {{ color: #ABABAB; }}

  /* ── Attendance badges ───────────────────────────────────── */
  .badge {{
    display: inline-block; padding: 3px 12px;
    border-radius: 12px; font-size: 12px; font-weight: 700;
    min-width: 50px; text-align: center;
  }}
  .badge-green  {{ background: #D4EDDA; color: #1A6B35; }}
  .badge-yellow {{ background: #FFF3CD; color: #856404; }}
  .badge-orange {{ background: #FFE8CC; color: #924800; }}
  .badge-red    {{ background: #FDDEDE; color: #A52020; }}

  /* ── Expand All button ───────────────────────────────────── */
  .toggle-btn {{
    background: none;
    border: 1.5px solid #CCAB44;
    color: #CCAB44;
    font-size: 11px; font-weight: 700;
    padding: 5px 14px; border-radius: 20px;
    cursor: pointer; letter-spacing: 0.5px;
    transition: background 0.15s, color 0.15s;
  }}
  .toggle-btn:hover {{ background: #F0B429; border-color: #F0B429; color: #fff; }}

  /* ── Manager accordion blocks ────────────────────────────── */
  .by-manager-grid {{ display: flex; flex-direction: column; gap: 10px; }}

  details.mgr-block {{
    border-radius: 7px;
    overflow: hidden;
    border: 1px solid #E4E0DA;
  }}

  details.mgr-block > summary {{
    list-style: none;
    cursor: pointer;
    background: #FBF8F2;
    border-left: 4px solid #F0B429;
    padding: 12px 18px;
    display: flex; align-items: center; justify-content: space-between;
    user-select: none;
    transition: background 0.15s;
  }}
  details.mgr-block > summary::-webkit-details-marker {{ display: none; }}
  details.mgr-block > summary:hover {{ background: #FFF4DC; }}

  .mgr-left {{ display: flex; align-items: center; gap: 10px; }}
  .mgr-name-title {{ font-size: 13px; font-weight: 700; color: #2D2D2D; }}

  .chevron {{
    color: #BBAC8A;
    font-size: 11px;
    display: inline-block;
    transition: transform 0.2s ease;
  }}
  details[open] > summary .chevron {{ transform: rotate(90deg); }}

  .mgr-meta {{
    font-size: 12px; color: #888;
    display: flex; align-items: center; gap: 8px;
  }}
  .risk-pill {{
    background: #FDDEDE; color: #A52020;
    font-size: 11px; font-weight: 700;
    padding: 2px 9px; border-radius: 10px;
    border: 1px solid #F0B0B0;
  }}

  .mgr-table-wrap {{ animation: slideDown 0.18s ease; }}
  @keyframes slideDown {{
    from {{ opacity: 0; transform: translateY(-4px); }}
    to   {{ opacity: 1; transform: translateY(0); }}
  }}

  /* ── Zero attendance section ─────────────────────────────── */
  .zero-section thead th {{ background: #7B2121; color: #fff; }}
  .zero-section .section-header h2::before {{ background: #C0392B; }}

  /* ── Footer ──────────────────────────────────────────────── */
  .footer {{
    background: #1A1A1C;
    padding: 13px 40px;
    display: flex; align-items: center; justify-content: space-between;
  }}
  .footer-logo {{ font-size: 13px; font-weight: 900; color: #F0B429; letter-spacing: 1px; }}
  .footer-note {{ font-size: 11px; color: rgba(255,255,255,0.3); }}
</style>
</head>
<body>
<div class="report">

  <!-- Header -->
  <div class="header">
    <div class="header-left">
      {logo_html}
    </div>
    <div class="header-right">
      <div class="report-title">Weekly Attendance Report</div>
      <div class="report-meta">
        <strong>{period}</strong> &nbsp;&bull;&nbsp;
        {total_weekdays} working day{"s" if total_weekdays != 1 else ""}
      </div>
    </div>
  </div>

  <!-- Stat cards -->
  <div class="cards">{cards}</div>

  <!-- All Employees -->
  <div class="section">
    <div class="section-header">
      <h2>All Employees</h2>
      <span class="pill">{total_emp} total</span>
    </div>
    {all_table}
  </div>

  <!-- By Manager -->
  <div class="section">
    <div class="section-header">
      <h2>By Manager</h2>
      <button class="toggle-btn" onclick="toggleAll(this)">Expand All</button>
    </div>
    <div class="by-manager-grid" id="mgr-grid">
      {mgr_sections}
    </div>
  </div>

  <!-- 0 Attendance -->
  {zero_section}

  <!-- Footer -->
  <div class="footer">
    <div class="footer-logo">TECHSUR</div>
    <div class="footer-note">Confidential &bull; For internal use only &bull; {generated_on}</div>
  </div>

</div>

<script>
  function toggleAll(btn) {{
    var blocks = document.querySelectorAll('#mgr-grid details');
    var anyCollapsed = Array.from(blocks).some(function(d) {{ return !d.open; }});
    blocks.forEach(function(d) {{ d.open = anyCollapsed; }});
    btn.textContent = anyCollapsed ? 'Collapse All' : 'Expand All';
  }}
</script>

</body>
</html>"""


# ── Sample data ────────────────────────────────────────────────────────────────

start          = date(2026, 3, 9)
end            = date(2026, 3, 13)
total_weekdays = 5

unique_days = pd.DataFrame([
    {"_name": "Alice Johnson",   "Days Present": 5, "Days Absent": 0, "Total Weekdays": 5, "Attendance %": 100.0, "Manager": "Craig Park",       "Manager Email": "craig.park@techsur.solutions"},
    {"_name": "Bob Martinez",    "Days Present": 4, "Days Absent": 1, "Total Weekdays": 5, "Attendance %":  80.0, "Manager": "Craig Park",       "Manager Email": "craig.park@techsur.solutions"},
    {"_name": "Carol White",     "Days Present": 3, "Days Absent": 2, "Total Weekdays": 5, "Attendance %":  60.0, "Manager": "Shailendra Gohil", "Manager Email": "shailendra.gohil@techsur.solutions"},
    {"_name": "David Brown",     "Days Present": 5, "Days Absent": 0, "Total Weekdays": 5, "Attendance %": 100.0, "Manager": "Shailendra Gohil", "Manager Email": "shailendra.gohil@techsur.solutions"},
    {"_name": "Eve Thompson",    "Days Present": 1, "Days Absent": 4, "Total Weekdays": 5, "Attendance %":  20.0, "Manager": "Craig Park",       "Manager Email": "craig.park@techsur.solutions"},
    {"_name": "Frank Wilson",    "Days Present": 2, "Days Absent": 3, "Total Weekdays": 5, "Attendance %":  40.0, "Manager": "Shailendra Gohil", "Manager Email": "shailendra.gohil@techsur.solutions"},
    {"_name": "Grace Lee",       "Days Present": 5, "Days Absent": 0, "Total Weekdays": 5, "Attendance %": 100.0, "Manager": "Craig Park",       "Manager Email": "craig.park@techsur.solutions"},
    {"_name": "Henry Garcia",    "Days Present": 4, "Days Absent": 1, "Total Weekdays": 5, "Attendance %":  80.0, "Manager": "Shailendra Gohil", "Manager Email": "shailendra.gohil@techsur.solutions"},
    {"_name": "Irene Robinson",  "Days Present": 3, "Days Absent": 2, "Total Weekdays": 5, "Attendance %":  60.0, "Manager": "Craig Park",       "Manager Email": "craig.park@techsur.solutions"},
    {"_name": "Jack Anderson",   "Days Present": 0, "Days Absent": 5, "Total Weekdays": 5, "Attendance %":   0.0, "Manager": "No Manager",       "Manager Email": ""},
    {"_name": "Karen Thomas",    "Days Present": 5, "Days Absent": 0, "Total Weekdays": 5, "Attendance %": 100.0, "Manager": "Shailendra Gohil", "Manager Email": "shailendra.gohil@techsur.solutions"},
    {"_name": "Leo Harris",      "Days Present": 4, "Days Absent": 1, "Total Weekdays": 5, "Attendance %":  80.0, "Manager": "Craig Park",       "Manager Email": "craig.park@techsur.solutions"},
    {"_name": "Mia Nguyen",      "Days Present": 2, "Days Absent": 3, "Total Weekdays": 5, "Attendance %":  40.0, "Manager": "Craig Park",       "Manager Email": "craig.park@techsur.solutions"},
    {"_name": "Noah Patel",      "Days Present": 5, "Days Absent": 0, "Total Weekdays": 5, "Attendance %": 100.0, "Manager": "Shailendra Gohil", "Manager Email": "shailendra.gohil@techsur.solutions"},
])

zero_df = pd.DataFrame([
    {"_name": "Mary Absent", "Days Present": 0, "Days Absent": 5, "Total Weekdays": 5, "Attendance %": 0.0, "Manager": "Craig Park",       "Manager Email": "craig.park@techsur.solutions"},
    {"_name": "Nathan Zero",  "Days Present": 0, "Days Absent": 5, "Total Weekdays": 5, "Attendance %": 0.0, "Manager": "Shailendra Gohil", "Manager Email": "shailendra.gohil@techsur.solutions"},
])

# ── Excel generator (mirrors weekly_report.py) ────────────────────────────────

def _safe_sheet_name(name):
    return (
        name[:31]
        .replace("/", "-").replace("\\", "-")
        .replace("*", "").replace("[", "").replace("]", "")
        .replace(":", "").replace("?", "")
    )


def _apply_sheet_formatting_xl(ws, df_cols, title, subtitle, tab_color="F0B429"):
    n_cols    = len(df_cols) + 1
    col_names = ["#"] + list(df_cols)

    ws.sheet_properties.tabColor = tab_color

    GOLD     = "F0B429"
    HDR_BG   = "3D3A35"
    EVEN_BG  = "FAFAF8"
    BORDER_C = "ECECEC"

    _gold_left  = Border(left=Side(style="thick", color=GOLD))
    _gold_btm   = Side(style="medium", color=GOLD)
    _row_border = Border(bottom=Side(style="thin", color=BORDER_C))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    tc = ws.cell(row=1, column=1)
    tc.value     = title
    tc.font      = Font(bold=True, size=14, color="2D2D2D", name="Calibri")
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    tc.border    = _gold_left

    sc = ws.cell(row=2, column=1)
    sc.value     = subtitle
    sc.font      = Font(italic=True, size=10, color="9E9E9E", name="Calibri")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sc.border    = _gold_left

    _NUM_COLS = {"#", "Days Present", "Days Absent", "Total Weekdays", "Attendance %"}
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

    for c_idx, col_name in enumerate(col_names, start=1):
        cell        = ws.cell(row=4, column=c_idx)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.border = Border(bottom=_gold_btm)
        is_num      = str(col_name) in _NUM_COLS
        cell.alignment = Alignment(
            horizontal="center" if is_num else "left",
            vertical="center", wrap_text=True,
            indent=0 if is_num else 1,
        )
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = ws.cell(row=5, column=1)

    pct_idx = next((i+1 for i, c in enumerate(col_names) if "Attendance" in str(c)), None)
    emp_idx = next((i+1 for i, c in enumerate(col_names) if str(c) == "Employee"),   None)

    even_fill = PatternFill("solid", fgColor=EVEN_BG)
    pct_fills = {
        "zero":    PatternFill("solid", fgColor="FDDEDE"),
        "atrisk":  PatternFill("solid", fgColor="FFE8CC"),
        "caution": PatternFill("solid", fgColor="FFF3CD"),
        "good":    PatternFill("solid", fgColor="D4EDDA"),
    }
    pct_fonts = {
        "zero":    Font(name="Calibri", size=10, bold=True, color="A52020"),
        "atrisk":  Font(name="Calibri", size=10, bold=True, color="924800"),
        "caution": Font(name="Calibri", size=10, bold=True, color="856404"),
        "good":    Font(name="Calibri", size=10, bold=True, color="1A6B35"),
    }
    emp_font  = Font(name="Calibri", size=10, bold=True, color="2D2D2D")
    num_font  = Font(name="Calibri", size=10,             color="4A4A4A")
    dim_font  = Font(name="Calibri", size=10,             color="ABABAB")
    data_font = Font(name="Calibri", size=10,             color="2D2D2D")

    for r_idx, row_cells in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=0):
        is_even = r_idx % 2 == 0
        for c_idx_0, cell in enumerate(row_cells):
            c_idx    = c_idx_0 + 1
            col_name = col_names[c_idx_0] if c_idx_0 < len(col_names) else ""
            is_num   = str(col_name) in _NUM_COLS and str(col_name) != "Attendance %"
            cell.border    = _row_border
            cell.alignment = Alignment(
                horizontal="center" if is_num else "left",
                vertical="center", indent=0 if is_num else 1,
            )
            if c_idx == pct_idx:
                pass
            elif c_idx == emp_idx:
                cell.font = emp_font
                if is_even: cell.fill = even_fill
            elif str(col_name) == "#":
                cell.font = dim_font
                if is_even: cell.fill = even_fill
            elif is_num:
                cell.font = num_font
                if is_even: cell.fill = even_fill
            else:
                cell.font = data_font
                if is_even: cell.fill = even_fill

        if pct_idx:
            pct_cell = row_cells[pct_idx - 1]
            try:
                val = float(pct_cell.value)
                key = ("zero" if val==0 else "atrisk" if val<80 else "caution" if val<100 else "good")
                pct_cell.fill          = pct_fills[key]
                pct_cell.font          = pct_fonts[key]
                pct_cell.number_format = '0.0"%"'
                pct_cell.alignment     = Alignment(horizontal="center", vertical="center")
                pct_cell.border        = _row_border
            except (TypeError, ValueError):
                pass

    col_widths = {
        "#": 4, "Employee": 24, "Days Present": 13, "Days Absent": 12,
        "Total Weekdays": 14, "Attendance %": 14, "Manager": 24, "Manager Email": 30,
    }
    for c_idx, col_name in enumerate(col_names, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(str(col_name), 15)


def _team_sheet_xl(df_team, writer, sheet_name, title, subtitle, tab_color="8B8680"):
    cols = ["_name", "Days Present", "Days Absent", "Total Weekdays", "Attendance %"]
    sheet_df = (
        df_team[cols]
        .sort_values("Attendance %", ascending=True)
        .rename(columns={"_name": "Employee"})
        .reset_index(drop=True)
    )
    sheet_df.index += 1
    sheet_df.to_excel(writer, sheet_name=sheet_name, index=True, index_label="#", startrow=3)
    _apply_sheet_formatting_xl(writer.sheets[sheet_name], list(sheet_df.columns), title, subtitle, tab_color)


def generate_report_excel(unique_days, zero_df, start, end):
    period = "{} \u2013 {}".format(start.strftime("%b %d"), end.strftime("%b %d, %Y"))
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # All Employees sheet
        cols = ["_name", "Days Present", "Days Absent", "Total Weekdays", "Attendance %"]
        if "Manager" in unique_days.columns:
            cols += ["Manager"]
        summary = (
            unique_days[cols]
            .sort_values(
                ["Manager", "Attendance %"] if "Manager" in cols else ["Attendance %"],
                ascending=True,
            )
            .rename(columns={"_name": "Employee"})
            .reset_index(drop=True)
        )
        summary.index += 1
        summary.to_excel(writer, sheet_name="All Employees", index=True, index_label="#", startrow=3)
        _apply_sheet_formatting_xl(
            writer.sheets["All Employees"], list(summary.columns),
            "TechSur Attendance Report",
            "Period: {}  |  All Employees".format(period),
            tab_color="F0B429",
        )

        if "Manager" in unique_days.columns:
            named_managers = sorted([
                m for m in unique_days["Manager"].dropna().unique()
                if m not in ("No Manager", "Unknown / Not Mapped")
            ])
            for mgr in named_managers:
                team = unique_days[unique_days["Manager"] == mgr].copy()
                if not team.empty:
                    _team_sheet_xl(
                        team, writer, _safe_sheet_name(mgr),
                        title=mgr, subtitle="Period: {}".format(period),
                        tab_color="5D7B8A",
                    )

            no_mgr = unique_days[
                unique_days["Manager"].isin(["No Manager", "Unknown / Not Mapped"])
            ].copy()
            if not no_mgr.empty:
                _team_sheet_xl(
                    no_mgr, writer, "No Manager",
                    title="No Manager Assigned", subtitle="Period: {}".format(period),
                    tab_color="AAAAAA",
                )

        if not zero_df.empty:
            _team_sheet_xl(
                zero_df, writer, "0 Attendance",
                title="0 Attendance \u2014 No Badge Swipes Recorded",
                subtitle="Period: {}".format(period),
                tab_color="C0392B",
            )

    return output.getvalue()


# ── Generate & open ────────────────────────────────────────────────────────────

html_content = generate_report_html(unique_days, zero_df, start, end, total_weekdays)
with open("sample_report.html", "w", encoding="utf-8") as f:
    f.write(html_content)
print("Saved: sample_report.html")
subprocess.run(["open", "sample_report.html"])

excel_bytes = generate_report_excel(unique_days, zero_df, start, end)
with open("sample_report.xlsx", "wb") as f:
    f.write(excel_bytes)
print("Saved: sample_report.xlsx")
subprocess.run(["open", "sample_report.xlsx"])
