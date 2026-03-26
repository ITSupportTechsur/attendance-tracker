"""
Weekly Attendance Automation
============================
Runs every Monday via GitHub Actions.
  1. Logs into D3000 Express (Datawatch DirectAccess) via browser automation
  2. Downloads the previous week's (Mon–Fri) badge-access Excel export
  3. Processes attendance using the same logic as attendance_app.py
  4. Generates the full multi-sheet Excel report
  5. Uploads the report to SharePoint IT Support Operations
  6. Emails the report (summary + Excel attachment) to the configured recipients

Required environment variables (set as GitHub Secrets):
  DATAWATCH_USERNAME    e.g. A.Admin5
  DATAWATCH_PASSWORD
  AZURE_TENANT_ID
  AZURE_CLIENT_ID
  AZURE_CLIENT_SECRET
  REPORT_FROM_EMAIL     mailbox to send from  e.g. Joe.ghaleb@techsur.solutions
  REPORT_TO_EMAILS      comma-separated recipients e.g. joe@techsur.solutions,manager@techsur.solutions
  TEAMS_WEBHOOK_URL     (optional) Power Automate webhook URL for Teams channel post
"""

import os
import io
import re
import sys
import base64
import difflib
import logging
from datetime import date, datetime, timedelta
from html import escape
from pathlib import Path

import pandas as pd
import msal
import requests as http_requests
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ── Configuration ──────────────────────────────────────────────────────────────

OFFICE_ADDRESS      = "11190 Sunrise Valley Drive"
TECHSUR_TENANT      = "Techsur Solutions"
DATAWATCH_BASE_URL  = "https://d3000express.azurewebsites.net"

DATAWATCH_USERNAME  = os.environ["DATAWATCH_USERNAME"]
DATAWATCH_PASSWORD  = os.environ["DATAWATCH_PASSWORD"]
AZURE_TENANT_ID     = os.environ["AZURE_TENANT_ID"]
AZURE_CLIENT_ID     = os.environ["AZURE_CLIENT_ID"]
AZURE_CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]
REPORT_FROM_EMAIL   = os.environ["REPORT_FROM_EMAIL"]   # mailbox to send from
REPORT_TO_EMAILS    = os.environ["REPORT_TO_EMAILS"]    # comma-separated recipients
TEAMS_WEBHOOK_URL   = os.environ.get("TEAMS_WEBHOOK_URL", "")  # optional

SHAREPOINT_SITE_PATH = "techsur.sharepoint.com:/sites/ITSupportOperations"
UPLOAD_FOLDER        = "Attendance Reports"   # folder inside site's document library

DEFAULT_EXCLUDE_NAMES = {
    "chief engineer master",
    "contractor shred it",
    "harvard 2",
    "guest fob 1",
    "guest fob 2",
    "ramona shannon harvard maintenance",
    "cristian mata",
    "bravo handy man",
    "rupinder yadav",       # removed from system per management
}

# Employees who are intentionally without a manager (e.g. company owner).
# They still appear in the report but are not flagged under "No Manager".
OWNER_EXCEPTIONS = {
    "amit yadav",           # company owner — no manager by design
}

_BADGE_JUNK_WORDS = {"lost", "spare", "inventory", "handy"}
_ISO_DATE_RE      = re.compile(r"^\d{4}-\d{2}-\d{2}T")

# Date format used by D3000 (e.g. 3/17/2026)
_DATE_FMT = "%#m/%#d/%Y" if sys.platform == "win32" else "%-m/%-d/%Y"


# ── Date helpers ───────────────────────────────────────────────────────────────

def get_last_week_range():
    """Return (last_monday, last_friday) of the most recently completed work week.
    Works correctly whether triggered on Monday (scheduled) or any other day (manual)."""
    today = date.today()
    # weekday(): Mon=0 Tue=1 Wed=2 Thu=3 Fri=4 Sat=5 Sun=6
    days_since_friday = (today.weekday() - 4) % 7  # 0=Fri,1=Sat,2=Sun,3=Mon,...
    if days_since_friday == 0:          # today is Friday — week not done yet
        days_since_friday = 7
    last_friday = today - timedelta(days=days_since_friday)
    last_monday = last_friday - timedelta(days=4)
    return last_monday, last_friday


def count_weekdays(start: date, end: date) -> int:
    return sum(
        1 for n in range((end - start).days + 1)
        if (start + timedelta(n)).weekday() < 5
    )


# ── Name normalisation (mirrors attendance_app.py) ────────────────────────────

def _name_key(name: str) -> str:
    tokens  = str(name).strip().lower().split()
    filtered = [t for t in tokens if len(t) > 1 and not t.isdigit()]
    deduped  = [filtered[i] for i in range(len(filtered))
                if i == 0 or filtered[i] != filtered[i - 1]]
    if len(deduped) >= 2:
        return deduped[0] + " " + deduped[-1]
    return " ".join(deduped) if deduped else name.strip().lower()


def _is_junk_badge_name(name: str) -> bool:
    return any(t in _BADGE_JUNK_WORDS for t in name.lower().split())


def _last_first_initial_match(k: str, candidates: list) -> str | None:
    parts_k = k.split()
    if len(parts_k) < 2:
        return None
    for c in candidates:
        parts_c = c.split()
        if (len(parts_c) >= 2
                and parts_k[-1] == parts_c[-1]
                and parts_k[0][0] == parts_c[0][0]):
            return c
    return None


# ── Step 1: Download badge Excel from D3000 ───────────────────────────────────

def download_badge_excel(start: date, end: date) -> bytes:
    """Use Playwright (headless Chromium) to log in and export the History Excel."""
    log.info(f"Downloading D3000 badge log  {start} → {end}")
    start_str = start.strftime(_DATE_FMT)
    end_str   = end.strftime(_DATE_FMT)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox"])
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()
        page.set_default_timeout(60_000)   # 60 s for all actions

        try:
            # ── Login step 1: load page and enter username ────────────────────
            log.info(f"Navigating to {DATAWATCH_BASE_URL}")
            page.goto(DATAWATCH_BASE_URL, wait_until="domcontentloaded", timeout=60_000)
            page.wait_for_timeout(2000)   # let JS render the form

            log.info(f"Page title: {page.title()!r}  URL: {page.url!r}")

            # Try multiple selector patterns for the username field
            username_selector = (
                "input#UserName, "
                "input[name='UserName'], "
                "input[name='username'], "
                "input[type='text']"
            )
            page.wait_for_selector(username_selector, timeout=30_000)
            page.fill(username_selector, DATAWATCH_USERNAME)
            log.info("Username entered")

            # Click Next — try value attr or visible button text
            page.click(
                "input[value='Next'], "
                "button:has-text('Next'), "
                "input[type='submit']"
            )
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(1500)

            # ── Login step 2: enter password ──────────────────────────────────
            log.info(f"Password page URL: {page.url!r}")
            page.wait_for_selector("input[name='Password'], input[type='password']", timeout=30_000)
            page.fill("input[name='Password'], input[type='password']", DATAWATCH_PASSWORD)
            page.click(
                "input[value='Log On'], "
                "button:has-text('Log On'), "
                "input[type='submit']"
            )
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(2000)
            log.info(f"After login URL: {page.url!r}")

        except Exception as exc:
            # Save a screenshot so we can see what the browser saw
            page.screenshot(path="/tmp/d3000_login_debug.png", full_page=True)
            log.error(f"Login failed — screenshot saved to /tmp/d3000_login_debug.png")
            raise RuntimeError(f"D3000 login failed: {exc}") from exc

        log.info("Logged in to D3000 DirectAccess")

        # ── Navigate to History ───────────────────────────────────────────────
        page.goto(f"{DATAWATCH_BASE_URL}/History/Index", wait_until="domcontentloaded")
        page.wait_for_timeout(3000)

        # Debug: log all input fields visible on the page
        inputs = page.evaluate("""
            () => Array.from(document.querySelectorAll('input')).map(
                el => ({id: el.id, name: el.name, type: el.type, value: el.value})
            )
        """)
        log.info(f"History page inputs: {inputs}")
        page.screenshot(path="/tmp/d3000_history.png", full_page=True)

        # ── Set date range via JavaScript (bypasses calendar picker) ──────────
        # Find the Begin/End inputs dynamically from what's on the page
        page.evaluate(f"""
            (function() {{
                // Try multiple strategies to find date inputs
                var inputs = document.querySelectorAll('input[type="text"], input[type="date"], input:not([type])');
                var dateInputs = Array.from(inputs).filter(function(el) {{
                    var id = (el.id || '').toLowerCase();
                    var name = (el.name || '').toLowerCase();
                    return id.includes('begin') || id.includes('start') ||
                           name.includes('begin') || name.includes('start');
                }});
                if (dateInputs.length > 0) {{
                    dateInputs[0].value = '{start_str}';
                    dateInputs[0].dispatchEvent(new Event('change', {{bubbles: true}}));
                    dateInputs[0].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}

                var endInputs = Array.from(inputs).filter(function(el) {{
                    var id = (el.id || '').toLowerCase();
                    var name = (el.name || '').toLowerCase();
                    return id.includes('end') || name.includes('end');
                }});
                if (endInputs.length > 0) {{
                    endInputs[0].value = '{end_str}';
                    endInputs[0].dispatchEvent(new Event('change', {{bubbles: true}}));
                    endInputs[0].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
            }})();
        """)
        page.wait_for_timeout(1000)

        # Verify the values were set
        set_values = page.evaluate("""
            () => Array.from(document.querySelectorAll('input')).map(
                el => ({id: el.id, name: el.name, value: el.value})
            ).filter(el => el.value)
        """)
        log.info(f"Input values after JS set: {set_values}")
        log.info(f"Date range set: {start_str} → {end_str}")

        # ── Click "Search By Tenant" ──────────────────────────────────────────
        # Find the search button dynamically
        search_btn = page.evaluate("""
            () => {
                var btns = Array.from(document.querySelectorAll('input[type="submit"], button'));
                var match = btns.find(b =>
                    (b.value || b.textContent || '').toLowerCase().includes('search')
                );
                return match ? (match.value || match.textContent) : null;
            }
        """)
        log.info(f"Search button found: {search_btn!r}")

        page.click(
            "input[value='Search By Tenant'], "
            "button:has-text('Search By Tenant'), "
            "input[type='submit']"
        )
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(4000)
        log.info("Search complete — downloading Export to Excel")

        # ── Export to Excel ───────────────────────────────────────────────────
        # Find the Export to Excel link
        export_link = page.evaluate("""
            () => {
                var links = Array.from(document.querySelectorAll('a'));
                var match = links.find(a => a.textContent.toLowerCase().includes('export to excel'));
                return match ? match.href : null;
            }
        """)
        log.info(f"Export link found: {export_link!r}")

        with page.expect_download(timeout=60_000) as dl_info:
            page.click("a:has-text('Export to Excel')")
        download = dl_info.value
        file_path = download.path()
        data = Path(file_path).read_bytes()
        browser.close()

    log.info(f"Downloaded badge log  ({len(data):,} bytes)")
    return data


# ── Step 2: Get Microsoft Graph token ─────────────────────────────────────────

def get_graph_token() -> str:
    app = msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{AZURE_TENANT_ID}",
        client_credential=AZURE_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" not in result:
        raise RuntimeError(f"MSAL token error: {result.get('error_description')}")
    return result["access_token"]


# ── Step 3: Fetch Azure AD users + managers ───────────────────────────────────

def fetch_manager_df(token: str) -> pd.DataFrame:
    headers = {"Authorization": f"Bearer {token}"}
    url = (
        "https://graph.microsoft.com/v1.0/users"
        "?$select=displayName,mail"
        "&$expand=manager($select=displayName,mail)"
        "&$top=999"
    )
    users = []
    while url:
        resp = http_requests.get(url, headers=headers)
        data = resp.json()
        if "error" in data:
            log.warning(f"Graph users error: {data['error']}")
            break
        users.extend(data.get("value", []))
        url = data.get("@odata.nextLink")

    rows = []
    for u in users:
        mgr = u.get("manager")
        rows.append({
            "Employee":      u.get("displayName", ""),
            "Manager":       mgr.get("displayName", "No Manager") if mgr else "No Manager",
            "Manager Email": mgr.get("mail", "") if mgr else "",
        })
    mgr_df = pd.DataFrame(rows)

    if not mgr_df.empty:
        mgr_df["_key"]     = mgr_df["Employee"].apply(_name_key)
        mgr_df["_has_mgr"] = mgr_df["Manager"].apply(
            lambda m: 0 if m not in ("No Manager", "") else 1
        )
        mgr_df = (
            mgr_df.sort_values("_has_mgr")
                  .drop_duplicates(subset=["_key"], keep="first")
                  .drop(columns=["_key", "_has_mgr"])
        )

    log.info(f"Fetched {len(mgr_df)} Azure AD users")
    return mgr_df


# ── Step 4: Fetch SharePoint site ID + DataWatch assignees ────────────────────

def get_sharepoint_site_id(token: str) -> str | None:
    """Resolve the SharePoint site path to a stable site ID."""
    headers = {"Authorization": f"Bearer {token}"}
    resp = http_requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE_PATH}",
        headers=headers,
    )
    data = resp.json()
    if "error" in data:
        log.warning(f"SharePoint site error: {data['error']}")
        return None
    return data["id"]


def fetch_datawatch_names(token: str, site_id: str) -> set:
    headers = {"Authorization": f"Bearer {token}"}

    lists_resp = http_requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists",
        headers=headers,
    )
    all_lists = lists_resp.json().get("value", [])
    list_id   = next(
        (l["id"] for l in all_lists if "hardware asset" in l.get("displayName", "").lower()),
        None,
    )
    if not list_id:
        log.warning("Hardware Asset Library not found in SharePoint — skipping 0-attendance check")
        return set()

    sp_url   = (
        f"https://graph.microsoft.com/v1.0/sites/{site_id}"
        f"/lists/{list_id}/items?$expand=fields&$top=999"
    )
    sp_items = []
    while sp_url:
        sp_resp = http_requests.get(sp_url, headers=headers)
        sp_data = sp_resp.json()
        if "error" in sp_data:
            log.warning(f"SharePoint items error: {sp_data['error']}")
            break
        sp_items.extend(sp_data.get("value", []))
        sp_url = sp_data.get("@odata.nextLink")

    names = set()
    for item in sp_items:
        fields = item.get("fields", {})
        if not any("datawatch" in str(v).lower() for v in fields.values()):
            continue
        raw = fields.get("field_1", "")
        if isinstance(raw, str):
            raw = raw.strip()
            if raw and not _ISO_DATE_RE.match(raw):
                names.add(raw)
        elif isinstance(raw, dict):
            n = raw.get("LookupValue") or raw.get("displayName") or ""
            if n:
                names.add(n.strip())

    log.info(f"Fetched {len(names)} DataWatch assignees from SharePoint")
    return names


# ── Step 5: Process attendance data ──────────────────────────────────────────

def find_col(df: pd.DataFrame, keywords: list) -> str | None:
    for col in df.columns:
        if any(k.lower() in col.lower() for k in keywords):
            return col
    return None


def process_attendance(
    excel_bytes: bytes,
    start: date,
    end: date,
    manager_df: pd.DataFrame,
    datawatch_names: set,
) -> tuple[pd.DataFrame, pd.DataFrame, int]:
    """
    Returns (unique_days_df, zero_attendance_df, total_weekdays).
    Mirrors the processing logic in attendance_app.py.
    """
    df_raw = pd.read_excel(io.BytesIO(excel_bytes))
    log.info(f"Badge log loaded: {len(df_raw):,} rows, columns: {list(df_raw.columns)}")

    # Auto-detect columns
    datetime_col  = find_col(df_raw, ["date", "time", "datetime", "timestamp"])
    firstname_col = find_col(df_raw, ["first"])
    lastname_col  = find_col(df_raw, ["last"])
    address_col   = find_col(df_raw, ["address", "from", "location", "site", "building"])
    tenant_col    = find_col(df_raw, ["tenant", "organization", "organisation", "company"])

    df = df_raw.copy()
    df["_dt"]   = pd.to_datetime(df[datetime_col], errors="coerce")
    df          = df.dropna(subset=["_dt"])
    df["_date"] = df["_dt"].dt.date

    if firstname_col and lastname_col:
        df["_name"] = (
            df[firstname_col].fillna("") + " " + df[lastname_col].fillna("")
        ).str.strip()
    else:
        nc = find_col(df_raw, ["name", "employee", "person", "user"])
        df["_name"] = df[nc].fillna("").str.strip() if nc else ""

    df = df[df["_name"] != ""]

    # Merge multiple fobs: "Craig Park 2" → "Craig Park" (strip trailing digit)
    df["_name"] = df["_name"].apply(
        lambda n: " ".join(n.split()[:-1]) if n.split() and n.split()[-1].isdigit() else n
    )
    df = df[df["_name"] != ""]

    # Address filter
    if address_col:
        df = df[df[address_col].astype(str).str.strip() == OFFICE_ADDRESS]

    # Tenant filter
    if tenant_col:
        df = df[df[tenant_col].astype(str).str.strip() == TECHSUR_TENANT]

    # Date range + weekdays only
    df = df[(df["_date"] >= start) & (df["_date"] <= end)]
    df = df[pd.to_datetime(df["_date"]).dt.dayofweek < 5]

    # Exclude default non-employee names
    df = df[~df["_name"].str.strip().str.lower().isin(DEFAULT_EXCLUDE_NAMES)]

    total_weekdays = count_weekdays(start, end)

    # Attendance per person
    unique_days = (
        df.drop_duplicates(subset=["_name", "_date"])
          .groupby("_name")["_date"]
          .count()
          .reset_index()
          .rename(columns={"_date": "Days Present"})
    )
    unique_days["Total Weekdays"] = total_weekdays
    unique_days["Attendance %"]   = (
        unique_days["Days Present"] / total_weekdays * 100
    ).round(1)
    unique_days["Days Absent"] = total_weekdays - unique_days["Days Present"]

    # Merge manager data
    if not manager_df.empty:
        mgr_lookup          = manager_df.copy()
        mgr_lookup["_key"]  = mgr_lookup["Employee"].apply(_name_key)
        unique_days["_key"] = unique_days["_name"].apply(_name_key)
        unique_days = unique_days.merge(
            mgr_lookup[["_key", "Manager", "Manager Email"]],
            on="_key", how="left",
        ).drop(columns=["_key"])
        unique_days["Manager"]       = unique_days["Manager"].fillna("Unknown / Not Mapped")
        unique_days["Manager Email"] = unique_days["Manager Email"].fillna("")

        # Owner exceptions: reclassify so they don't appear in "No Manager" section
        def _reclassify_owner(row):
            if (row["Manager"] in ("No Manager", "Unknown / Not Mapped")
                    and _name_key(row["_name"]) in OWNER_EXCEPTIONS):
                return "Owner / Executive"
            return row["Manager"]
        unique_days["Manager"] = unique_days.apply(_reclassify_owner, axis=1)

    # Zero-attendance: DataWatch holders with no badge swipes
    zero_rows = []
    if datawatch_names:
        existing_keys     = list(unique_days["_name"].apply(_name_key))
        existing_keys_set = set(existing_keys)
        for n in sorted(datawatch_names):
            if not n.strip():
                continue
            if n.strip().lower() in DEFAULT_EXCLUDE_NAMES or _is_junk_badge_name(n):
                continue
            k     = _name_key(n)
            if k in existing_keys_set:
                continue
            close = difflib.get_close_matches(k, existing_keys, n=1, cutoff=0.82)
            if not close:
                m = _last_first_initial_match(k, existing_keys)
                if m:
                    close = [m]
            if close:
                continue   # fuzzy match — person was present
            zero_rows.append({
                "_name": n, "Days Present": 0,
                "Days Absent": total_weekdays,
                "Total Weekdays": total_weekdays,
                "Attendance %": 0.0,
            })

    zero_df = pd.DataFrame(zero_rows)
    if not zero_df.empty and not manager_df.empty:
        mgr_lookup_z          = manager_df.copy()
        mgr_lookup_z["_key"]  = mgr_lookup_z["Employee"].apply(_name_key)
        zero_df["_key"]       = zero_df["_name"].apply(_name_key)
        zero_df = zero_df.merge(
            mgr_lookup_z[["_key", "Manager", "Manager Email"]],
            on="_key", how="left",
        ).drop(columns=["_key"])
        zero_df["Manager"]       = zero_df["Manager"].fillna("Unknown / Not Mapped")
        zero_df["Manager Email"] = zero_df["Manager Email"].fillna("")

    log.info(
        f"Processed: {len(unique_days)} employees, "
        f"{len(zero_df)} zero-attendance, "
        f"{total_weekdays} weekdays"
    )
    return unique_days, zero_df, total_weekdays


# ── Step 6: Generate Excel report ─────────────────────────────────────────────

def _safe_sheet_name(name: str) -> str:
    return (
        name[:31]
        .replace("/", "-").replace("\\", "-")
        .replace("*", "").replace("[", "").replace("]", "")
        .replace(":", "").replace("?", "")
    )


def _apply_sheet_formatting(
    ws, df_cols: list, title: str, subtitle: str, tab_color: str = "F0B429"
) -> None:
    """Apply professional formatting to a worksheet written with startrow=3."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    n_cols    = len(df_cols) + 1   # +1 for the index "#" column
    col_names = ["#"] + list(df_cols)

    # Sheet tab colour
    ws.sheet_properties.tabColor = tab_color

    # ── Palette ────────────────────────────────────────────────────────────
    GOLD     = "F0B429"
    HDR_BG   = "3D3A35"   # warm charcoal
    EVEN_BG  = "FAFAF8"   # warm off-white
    BORDER_C = "ECECEC"   # subtle row separator

    _gold_left  = Border(left=Side(style="thick", color=GOLD))
    _gold_btm   = Side(style="medium", color=GOLD)
    _row_border = Border(bottom=Side(style="thin", color=BORDER_C))

    # ── Title (row 1) & subtitle (row 2) ──────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    tc = ws.cell(row=1, column=1)
    tc.value     = title
    tc.font      = Font(bold=True, size=14, color="2D2D2D", name="Calibri")
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    tc.border    = _gold_left

    sc = ws.cell(row=2, column=1)
    sc.value     = subtitle
    sc.font      = Font(italic=True, size=10, color="9E9E9E", name="Calibri")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sc.border    = _gold_left

    # ── Header row (row 4) ─────────────────────────────────────────────────
    _NUM_COLS = {"#", "Days Present", "Days Absent", "Total Weekdays", "Attendance %"}
    hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

    for c_idx, col_name in enumerate(col_names, start=1):
        cell           = ws.cell(row=4, column=c_idx)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.border    = Border(bottom=_gold_btm)
        is_num         = str(col_name) in _NUM_COLS
        cell.alignment = Alignment(
            horizontal="center" if is_num else "left",
            vertical="center", wrap_text=True,
            indent=0 if is_num else 1,
        )
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = ws.cell(row=5, column=1)

    # ── Column indices ─────────────────────────────────────────────────────
    pct_idx = next(
        (i + 1 for i, c in enumerate(col_names) if "Attendance" in str(c)), None
    )
    emp_idx = next(
        (i + 1 for i, c in enumerate(col_names) if str(c) == "Employee"), None
    )

    # ── Fonts & fills ──────────────────────────────────────────────────────
    even_fill  = PatternFill("solid", fgColor=EVEN_BG)
    pct_fills  = {
        "zero":    PatternFill("solid", fgColor="FDDEDE"),
        "atrisk":  PatternFill("solid", fgColor="FFE8CC"),
        "caution": PatternFill("solid", fgColor="FFF3CD"),
        "good":    PatternFill("solid", fgColor="D4EDDA"),
    }
    pct_fonts  = {
        "zero":    Font(name="Calibri", size=10, bold=True, color="A52020"),
        "atrisk":  Font(name="Calibri", size=10, bold=True, color="924800"),
        "caution": Font(name="Calibri", size=10, bold=True, color="856404"),
        "good":    Font(name="Calibri", size=10, bold=True, color="1A6B35"),
    }
    emp_font   = Font(name="Calibri", size=10, bold=True,  color="2D2D2D")
    num_font   = Font(name="Calibri", size=10,             color="4A4A4A")
    dim_font   = Font(name="Calibri", size=10,             color="ABABAB")
    data_font  = Font(name="Calibri", size=10,             color="2D2D2D")

    # ── Data rows ──────────────────────────────────────────────────────────
    for r_idx, row_cells in enumerate(
        ws.iter_rows(min_row=5, max_row=ws.max_row), start=0
    ):
        is_even = r_idx % 2 == 0

        for c_idx_0, cell in enumerate(row_cells):
            c_idx    = c_idx_0 + 1
            col_name = col_names[c_idx_0] if c_idx_0 < len(col_names) else ""
            is_num   = str(col_name) in _NUM_COLS and str(col_name) != "Attendance %"

            cell.border    = _row_border
            cell.alignment = Alignment(
                horizontal="center" if is_num else "left",
                vertical="center",
                indent=0 if is_num else 1,
            )

            if c_idx == pct_idx:
                pass   # handled separately
            elif c_idx == emp_idx:
                cell.font = emp_font
                if is_even: cell.fill = even_fill
            elif str(col_name) == "#":
                cell.font = dim_font
                if is_even: cell.fill = even_fill
            elif is_num:
                cell.font = num_font
                if is_even: cell.fill = even_fill
            else:
                cell.font = data_font
                if is_even: cell.fill = even_fill

        if pct_idx:
            pct_cell = row_cells[pct_idx - 1]
            try:
                val = float(pct_cell.value)
                key = (
                    "zero"    if val == 0  else
                    "atrisk"  if val < 60  else
                    "caution" if val < 100 else
                    "good"
                )
                pct_cell.fill          = pct_fills[key]
                pct_cell.font          = pct_fonts[key]
                pct_cell.number_format = '0.0"%"'
                pct_cell.alignment     = Alignment(horizontal="center", vertical="center")
                pct_cell.border        = _row_border
            except (TypeError, ValueError):
                pass

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = {
        "#": 4, "Employee": 24, "Days Present": 13, "Days Absent": 12,
        "Total Weekdays": 14, "Attendance %": 14, "Manager": 24, "Manager Email": 30,
    }
    for c_idx, col_name in enumerate(col_names, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(
            str(col_name), 15
        )


def _team_sheet(
    df_team: pd.DataFrame,
    writer,
    sheet_name: str,
    title: str,
    subtitle: str,
    tab_color: str = "8B8680",
) -> None:
    cols = ["_name", "Days Present", "Days Absent", "Total Weekdays", "Attendance %"]
    sheet_df = (
        df_team[cols]
        .sort_values("Attendance %", ascending=True)
        .rename(columns={"_name": "Employee"})
        .reset_index(drop=True)
    )
    sheet_df.index += 1
    sheet_df.to_excel(
        writer, sheet_name=sheet_name,
        index=True, index_label="#", startrow=3,
    )
    _apply_sheet_formatting(
        writer.sheets[sheet_name], list(sheet_df.columns), title, subtitle, tab_color
    )


def generate_report_excel(
    unique_days: pd.DataFrame,
    zero_df: pd.DataFrame,
    start: date,
    end: date,
) -> bytes:
    period = f"{start.strftime('%b %d')} \u2013 {end.strftime('%b %d, %Y')}"
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # Sheet 1 — All Employees
        cols = ["_name", "Days Present", "Days Absent", "Total Weekdays", "Attendance %"]
        if "Manager" in unique_days.columns:
            cols += ["Manager"]
        summary = (
            unique_days[cols]
            .sort_values(
                ["Manager", "Attendance %"] if "Manager" in cols else ["Attendance %"],
                ascending=True,
            )
            .rename(columns={"_name": "Employee"})
            .reset_index(drop=True)
        )
        summary.index += 1
        summary.to_excel(
            writer, sheet_name="All Employees",
            index=True, index_label="#", startrow=3,
        )
        _apply_sheet_formatting(
            writer.sheets["All Employees"],
            list(summary.columns),
            "TechSur Attendance Report",
            f"Period: {period}  |  All Employees",
            tab_color="F0B429",   # gold — primary sheet
        )

        # One sheet per manager
        if "Manager" in unique_days.columns:
            named_managers = sorted(
                [m for m in unique_days["Manager"].dropna().unique()
                 if m not in ("No Manager", "Unknown / Not Mapped")],
                key=lambda m: (0 if m == "Owner / Executive" else 1, m)
            )
            for mgr in named_managers:
                team = unique_days[unique_days["Manager"] == mgr].copy()
                if not team.empty:
                    _team_sheet(
                        team, writer, _safe_sheet_name(mgr),
                        title=mgr,
                        subtitle=f"Period: {period}",
                        tab_color="5D7B8A",   # muted slate blue
                    )

            no_mgr = unique_days[
                unique_days["Manager"].isin(["No Manager", "Unknown / Not Mapped"])
            ].copy()
            if not no_mgr.empty:
                _team_sheet(
                    no_mgr, writer, "No Manager",
                    title="No Manager Assigned",
                    subtitle=f"Period: {period}",
                    tab_color="AAAAAA",   # grey
                )

        # 0 Attendance sheet
        if not zero_df.empty:
            _team_sheet(
                zero_df, writer, "0 Attendance",
                title="0 Attendance \u2014 No Badge Swipes Recorded",
                subtitle=f"Period: {period}",
                tab_color="C0392B",   # red
            )

    log.info("Excel report generated")
    return output.getvalue()


# ── Step 6b: Generate HTML report ─────────────────────────────────────────────

def _html_pct_badge(val: float) -> str:
    try:
        v = float(val)
        if v == 0:    cls, label = "badge-red",    f"{v:.0f}%"
        elif v < 60:  cls, label = "badge-orange",  f"{v:.0f}%"
        elif v < 100: cls, label = "badge-yellow",  f"{v:.0f}%"
        else:         cls, label = "badge-green",   f"{v:.0f}%"
        return f'<span class="badge {cls}">{label}</span>'
    except (TypeError, ValueError):
        return escape(str(val))


def _html_table(df: pd.DataFrame, show_manager: bool = True) -> str:
    has_mgr = show_manager and "Manager" in df.columns
    mgr_th  = "<th style='text-align:left'>Manager</th>" if has_mgr else ""
    rows = ""
    for i, (_, row) in enumerate(df.iterrows(), 1):
        mgr_cell = (
            f'<td class="muted">{escape(str(row.get("Manager", "")))}</td>'
            if has_mgr else ""
        )
        rows += (
            f"<tr>"
            f'<td class="num light">{i}</td>'
            f'<td class="emp">{escape(str(row["_name"]))}</td>'
            f"{mgr_cell}"
            f'<td class="num">{int(row["Days Present"])}</td>'
            f'<td class="num">{int(row["Days Absent"])}</td>'
            f'<td class="num light">{int(row["Total Weekdays"])}</td>'
            f'<td class="center">{_html_pct_badge(row["Attendance %"])}</td>'
            f"</tr>"
        )
    return (
        f'<table><thead><tr>'
        f'<th style="width:36px">#</th><th style="text-align:left">Employee</th>'
        f"{mgr_th}"
        f'<th>Present</th><th>Absent</th><th>Total</th><th>Attendance</th>'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )


def generate_report_html(
    unique_days: pd.DataFrame,
    zero_df: pd.DataFrame,
    start: date,
    end: date,
    total_weekdays: int,
) -> bytes:
    period       = f"{start.strftime('%B %d')} \u2013 {end.strftime('%B %d, %Y')}"
    generated_on = datetime.now().strftime("%B %d, %Y")

    total_emp  = len(unique_days)
    avg_pct    = unique_days["Attendance %"].mean() if total_emp else 0.0
    at_risk    = int((unique_days["Attendance %"] < 60).sum())
    zero_count = len(zero_df)

    # Logo (base64-embedded if file exists)
    logo_path = Path(__file__).parent / "techsur_logo.png"
    if logo_path.exists():
        logo_html = (
            f'<img src="data:image/png;base64,'
            f'{base64.b64encode(logo_path.read_bytes()).decode()}" '
            f'alt="TechSur" style="height:52px;display:block;">'
        )
    else:
        logo_html = (
            '<div style="display:flex;flex-direction:column;align-items:flex-start;">'
            '<div style="font-size:28px;font-weight:900;color:#F0B429;letter-spacing:2px;'
            'font-family:Arial,sans-serif;line-height:1;">TECHSUR</div>'
            '<div style="font-size:8px;color:rgba(255,255,255,0.55);letter-spacing:2.5px;'
            'font-weight:600;text-transform:uppercase;margin-top:3px;">PASSION MEETS TECHNOLOGY</div>'
            '</div>'
        )

    # All Employees table
    sort_cols  = ["Manager", "Attendance %"] if "Manager" in unique_days.columns else ["Attendance %"]
    all_sorted = unique_days.sort_values(sort_cols, ascending=True)
    all_table  = _html_table(all_sorted, show_manager=True)

    # By-Manager collapsible sections
    mgr_sections = ""
    if "Manager" in unique_days.columns:
        named = sorted(
            [m for m in unique_days["Manager"].dropna().unique()
             if m not in ("No Manager", "Unknown / Not Mapped", "")],
            key=lambda m: (0 if m == "Owner / Executive" else 1, m)
        )
        for mgr in named:
            team   = unique_days[unique_days["Manager"] == mgr].copy()
            avg    = team["Attendance %"].mean()
            n      = len(team)
            risk_n = int((team["Attendance %"] < 60).sum())
            tbl    = _html_table(team.sort_values("Attendance %"), show_manager=False)
            risk_pill = f'<span class="risk-pill">{risk_n} at risk</span>' if risk_n > 0 else ""
            mgr_sections += (
                f'<details class="mgr-block">'
                f'<summary class="mgr-bar">'
                f'<span class="mgr-left"><span class="chevron">&#9656;</span>'
                f'<span class="mgr-name-title">{escape(mgr)}</span></span>'
                f'<span class="mgr-meta">{n} employee{"s" if n!=1 else ""}'
                f'&nbsp;&nbsp;&bull;&nbsp;&nbsp;Avg: {_html_pct_badge(avg)}&nbsp;&nbsp;{risk_pill}'
                f'</span></summary>'
                f'<div class="mgr-table-wrap">{tbl}</div></details>'
            )
        no_mgr = unique_days[
            unique_days["Manager"].isin(["No Manager", "Unknown / Not Mapped"])
        ].copy()
        if not no_mgr.empty:
            tbl = _html_table(no_mgr.sort_values("Attendance %"), show_manager=False)
            mgr_sections += (
                f'<details class="mgr-block">'
                f'<summary class="mgr-bar" style="border-left-color:#555;">'
                f'<span class="mgr-left"><span class="chevron">&#9656;</span>'
                f'<span class="mgr-name-title" style="color:#aaa;">No Manager Assigned</span></span>'
                f'<span class="mgr-meta">{len(no_mgr)} employee{"s" if len(no_mgr)!=1 else ""}</span>'
                f'</summary>'
                f'<div class="mgr-table-wrap">{tbl}</div></details>'
            )

    # 0 Attendance section
    zero_section = ""
    if not zero_df.empty:
        zero_tbl = _html_table(zero_df.sort_values("_name"), show_manager=True)
        zero_section = (
            f'<div class="section zero-section">'
            f'<div class="section-header">'
            f'<h2 style="color:#C0392B;">&#9888; Zero Attendance</h2>'
            f'<span class="pill" style="background:#FDDEDE;color:#A52020;">{len(zero_df)}</span>'
            f'</div>'
            f'<p class="note">These employees have a DataWatch badge assigned but no recorded'
            f' office entries for this period.</p>'
            f'{zero_tbl}</div>'
        )

    # Stat cards
    avg_color  = "#27AE60" if avg_pct >= 60 else "#E67E22" if avg_pct >= 40 else "#E74C3C"
    risk_color = "#E74C3C" if at_risk    > 0 else "#27AE60"
    zero_color = "#E74C3C" if zero_count > 0 else "#27AE60"
    cards = (
        f'<div class="card"><div class="stat">{total_emp}</div>'
        f'<div class="stat-label">Hybrid Employees</div></div>'
        f'<div class="card"><div class="stat" style="color:{avg_color}">{avg_pct:.1f}%</div>'
        f'<div class="stat-label">Avg Attendance</div></div>'
        f'<div class="card"><div class="stat" style="color:{risk_color}">{at_risk}</div>'
        f'<div class="stat-label">At Risk (&lt;60%)</div></div>'
        f'<div class="card"><div class="stat" style="color:{zero_color}">{zero_count}</div>'
        f'<div class="stat-label">Zero Attendance</div></div>'
    )

    css = """
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    body{font-family:'Segoe UI','Helvetica Neue',Arial,sans-serif;background:#F0F2F5;color:#2D2D2D;padding:28px 16px;font-size:14px;line-height:1.5}
    .report{max-width:980px;margin:0 auto;background:#fff;border-radius:10px;box-shadow:0 2px 16px rgba(0,0,0,.09);overflow:hidden}
    .header{background:#1A1A1C;padding:24px 40px;display:flex;align-items:center;justify-content:space-between;border-bottom:3px solid #F0B429}
    .header-right{text-align:right}
    .header-right .report-title{font-size:19px;font-weight:700;color:#fff;margin-bottom:4px}
    .header-right .report-meta{font-size:12px;color:rgba(255,255,255,.5)}
    .header-right .report-meta strong{color:#F0B429;font-weight:600}
    .cards{display:flex;background:#FAFAF8;border-bottom:1px solid #ECECEC}
    .card{flex:1;padding:20px 16px;text-align:center;border-right:1px solid #ECECEC}
    .card:last-child{border-right:none}
    .stat{font-size:32px;font-weight:700;color:#2D2D2D;line-height:1;margin-bottom:5px}
    .stat-label{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.8px;color:#9E9E9E}
    .section{padding:26px 40px 34px;border-bottom:1px solid #ECECEC}
    .section:last-child{border-bottom:none}
    .section-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px}
    .section-header h2{font-size:15px;font-weight:700;color:#2D2D2D;display:flex;align-items:center;gap:10px}
    .section-header h2::before{content:'';display:inline-block;width:4px;height:18px;border-radius:2px;background:#F0B429}
    .pill{background:#FFF3CD;color:#856404;font-size:11px;font-weight:700;padding:3px 10px;border-radius:12px}
    .note{font-size:12px;color:#9E9E9E;margin-bottom:12px}
    .toggle-btn{background:none;border:1.5px solid #CCAB44;color:#CCAB44;font-size:11px;font-weight:700;padding:5px 14px;border-radius:20px;cursor:pointer;letter-spacing:.5px;transition:background .15s,color .15s}
    .toggle-btn:hover{background:#F0B429;border-color:#F0B429;color:#fff}
    table{width:100%;border-collapse:collapse;font-size:13px}
    thead th{background:#3D3A35;color:#fff;padding:10px 14px;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;text-align:center;white-space:nowrap}
    thead th[style*="left"]{text-align:left}
    tbody tr:nth-child(even) td{background:#FAFAF8}
    tbody tr:hover td{background:#FFF8E8!important}
    tbody td{padding:9px 14px;border-bottom:1px solid #F0EFED;vertical-align:middle}
    td.num{text-align:center;font-variant-numeric:tabular-nums;color:#4A4A4A}
    td.center{text-align:center}
    td.emp{font-weight:600;color:#2D2D2D}
    td.muted{color:#888;font-size:12px}
    td.light{color:#ABABAB}
    .badge{display:inline-block;padding:3px 12px;border-radius:12px;font-size:12px;font-weight:700;min-width:50px;text-align:center}
    .badge-green{background:#D4EDDA;color:#1A6B35}
    .badge-yellow{background:#FFF3CD;color:#856404}
    .badge-orange{background:#FFE8CC;color:#924800}
    .badge-red{background:#FDDEDE;color:#A52020}
    .by-manager-grid{display:flex;flex-direction:column;gap:10px}
    details.mgr-block{border-radius:7px;overflow:hidden;border:1px solid #E4E0DA}
    details.mgr-block>summary{list-style:none;cursor:pointer;background:#FBF8F2;border-left:4px solid #F0B429;padding:12px 18px;display:flex;align-items:center;justify-content:space-between;user-select:none;transition:background .15s}
    details.mgr-block>summary::-webkit-details-marker{display:none}
    details.mgr-block>summary:hover{background:#FFF4DC}
    .mgr-left{display:flex;align-items:center;gap:10px}
    .mgr-name-title{font-size:13px;font-weight:700;color:#2D2D2D}
    .chevron{color:#BBAC8A;font-size:11px;display:inline-block;transition:transform .2s ease}
    details[open]>summary .chevron{transform:rotate(90deg)}
    .mgr-meta{font-size:12px;color:#888;display:flex;align-items:center;gap:8px}
    .risk-pill{background:#FDDEDE;color:#A52020;font-size:11px;font-weight:700;padding:2px 9px;border-radius:10px;border:1px solid #F0B0B0}
    .mgr-table-wrap{animation:slideDown .18s ease}
    @keyframes slideDown{from{opacity:0;transform:translateY(-4px)}to{opacity:1;transform:translateY(0)}}
    .zero-section thead th{background:#7B2121;color:#fff}
    .zero-section .section-header h2::before{background:#C0392B}
    .footer{background:#1A1A1C;padding:13px 40px;display:flex;align-items:center;justify-content:space-between}
    .footer-logo{font-size:13px;font-weight:900;color:#F0B429;letter-spacing:1px}
    .footer-note{font-size:11px;color:rgba(255,255,255,.3)}
    """

    html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>TechSur Attendance Report \u2013 {period}</title>
<style>{css}</style></head>
<body><div class="report">
  <div class="header">
    <div class="header-left">{logo_html}</div>
    <div class="header-right">
      <div class="report-title">Weekly Attendance Report</div>
      <div class="report-meta"><strong>{period}</strong> &nbsp;&bull;&nbsp; {total_weekdays} working day{"s" if total_weekdays!=1 else ""}</div>
    </div>
  </div>
  <div class="cards">{cards}</div>
  <div class="section">
    <div class="section-header"><h2>All Employees</h2><span class="pill">{total_emp} total</span></div>
    {all_table}
  </div>
  <div class="section">
    <div class="section-header">
      <h2>By Manager</h2>
      <button class="toggle-btn" onclick="toggleAll(this)">Expand All</button>
    </div>
    <div class="by-manager-grid" id="mgr-grid">{mgr_sections}</div>
  </div>
  {zero_section}
  <div class="footer">
    <div class="footer-logo">TECHSUR</div>
    <div class="footer-note">Confidential &bull; For internal use only &bull; {generated_on}</div>
  </div>
</div>
<script>
  function toggleAll(btn){{
    var blocks=document.querySelectorAll('#mgr-grid details');
    var anyCollapsed=Array.from(blocks).some(function(d){{return !d.open;}});
    blocks.forEach(function(d){{d.open=anyCollapsed;}});
    btn.textContent=anyCollapsed?'Collapse All':'Expand All';
  }}
</script>
</body></html>"""

    log.info("HTML report generated")
    return html.encode("utf-8")


# ── Step 7: Upload file to SharePoint ─────────────────────────────────────────

def upload_to_sharepoint(
    token: str, site_id: str, filename: str, file_bytes: bytes,
    content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    week_folder: str = "",
) -> str:
    """
    Uploads the report to the SharePoint site's 'Attendance Reports' folder.
    If week_folder is given, files are placed inside a subfolder named after the audit week.
    Uses the resolved site_id (GUID) so the URL is unambiguous.
    Returns the web URL of the uploaded file, or "" on failure.
    Requires Sites.ReadWrite.All on the Azure AD app.
    """
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": content_type,
    }
    subfolder_path = f"{UPLOAD_FOLDER}/{week_folder}" if week_folder else UPLOAD_FOLDER
    upload_url = (
        f"https://graph.microsoft.com/v1.0/sites/{site_id}"
        f"/drive/root:/{subfolder_path}/{filename}:/content"
    )
    resp = http_requests.put(upload_url, headers=headers, data=file_bytes)
    if resp.status_code not in (200, 201):
        log.warning(f"SharePoint upload failed ({resp.status_code}): {resp.text[:200]}")
        return ""

    web_url = resp.json().get("webUrl", "")
    log.info(f"File uploaded to SharePoint: {web_url}")
    return web_url


# ── Step 8: Send email report ──────────────────────────────────────────────────

def send_email_report(
    token: str,
    from_email: str,
    to_emails: str,
    unique_days: pd.DataFrame,
    zero_df: pd.DataFrame,
    total_weekdays: int,
    start: date,
    end: date,
    file_url: str,
    filename: str,
    report_bytes: bytes,
    html_bytes: bytes | None = None,
    html_filename: str = "",
) -> None:
    """
    Sends the attendance report by email with Excel + HTML attachments.
    Uses Microsoft Graph sendMail — requires Mail.Send application permission.
    from_email   : mailbox to send from (must exist in the tenant)
    to_emails    : comma-separated list of recipient addresses
    """

    total_emp  = len(unique_days)
    avg_pct    = unique_days["Attendance %"].mean() if total_emp else 0.0
    at_risk    = int((unique_days["Attendance %"] < 60).sum())
    zero_count = len(zero_df)

    sharepoint_line = (
        f'<p>📎 Full report also saved to SharePoint: '
        f'<a href="{file_url}">{filename}</a></p>'
        if file_url else ""
    )

    html_body = f"""
<p>Please find attached the weekly attendance tracker results.</p>

<table style="border-collapse:collapse; font-family:Arial,sans-serif; font-size:14px;">
  <tr><td style="padding:4px 12px 4px 0;"><b>Period</b></td>
      <td>{start.strftime('%b %d')} – {end.strftime('%b %d, %Y')}</td></tr>
  <tr><td style="padding:4px 12px 4px 0;"><b>Working days</b></td>
      <td>{total_weekdays}</td></tr>
  <tr><td style="padding:4px 12px 4px 0;"><b>Employees tracked</b></td>
      <td>{total_emp}</td></tr>
  <tr><td style="padding:4px 12px 4px 0;"><b>Average attendance</b></td>
      <td>{avg_pct:.1f}%</td></tr>
  <tr><td style="padding:4px 12px 4px 0;"><b>At risk (&lt;60%)</b></td>
      <td>{at_risk}</td></tr>
  <tr><td style="padding:4px 12px 4px 0;"><b>0 attendance</b></td>
      <td>{zero_count}</td></tr>
</table>

{sharepoint_line}

<p style="color:#888;font-size:12px;">Sent automatically by the TechSur Attendance Tracker.</p>
"""

    recipients = [
        {"emailAddress": {"address": addr.strip()}}
        for addr in to_emails.split(",") if addr.strip()
    ]

    payload = {
        "message": {
            "subject": (
                f"Weekly Attendance Report — "
                f"{start.strftime('%b %d')} to {end.strftime('%b %d, %Y')}"
            ),
            "body": {"contentType": "HTML", "content": html_body},
            "toRecipients": recipients,
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": filename,
                    "contentType": (
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    ),
                    "contentBytes": base64.b64encode(report_bytes).decode(),
                },
                *([{
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": html_filename,
                    "contentType": "text/html",
                    "contentBytes": base64.b64encode(html_bytes).decode(),
                }] if html_bytes and html_filename else []),
            ],
        },
        "saveToSentItems": False,
    }

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    resp = http_requests.post(
        f"https://graph.microsoft.com/v1.0/users/{from_email}/sendMail",
        headers=headers,
        json=payload,
    )
    if resp.status_code == 202:
        log.info(f"Email sent from {from_email} to {to_emails}")
    else:
        raise RuntimeError(
            f"Email send failed ({resp.status_code}): {resp.text[:300]}"
        )


# ── Step 9: Post summary to Teams channel (via Power Automate webhook) ─────────

def post_to_teams_webhook(
    unique_days: pd.DataFrame,
    zero_df: pd.DataFrame,
    total_weekdays: int,
    start: date,
    end: date,
    file_url: str,
    html_url: str = "",
) -> None:
    if not TEAMS_WEBHOOK_URL:
        log.info("TEAMS_WEBHOOK_URL not set — skipping Teams post")
        return

    total_emp  = len(unique_days)
    avg_pct    = unique_days["Attendance %"].mean() if total_emp else 0.0
    at_risk    = int((unique_days["Attendance %"] < 60).sum())
    zero_count = len(zero_df)

    facts = [
        {"title": "Period",               "value": f"{start.strftime('%b %d')} – {end.strftime('%b %d, %Y')}"},
        {"title": "Working days",         "value": str(total_weekdays)},
        {"title": "Employees tracked",    "value": str(total_emp)},
        {"title": "Average attendance",   "value": f"{avg_pct:.1f}%"},
        {"title": "At risk (<60%)",       "value": str(at_risk)},
        {"title": "0 attendance",         "value": str(zero_count)},
    ]

    body = [
        {
            "type": "TextBlock",
            "size": "Large",
            "weight": "Bolder",
            "text": f"Weekly Attendance Report — {start.strftime('%b %d')} to {end.strftime('%b %d, %Y')}",
            "wrap": True,
        },
        {"type": "FactSet", "facts": facts},
    ]

    actions = []
    if html_url:
        actions.append({
            "type": "Action.OpenUrl",
            "title": "Open HTML Report",
            "url": html_url,
        })
    if file_url:
        actions.append({
            "type": "Action.OpenUrl",
            "title": "Download Excel Report",
            "url": file_url,
        })
    if actions:
        body.append({"type": "ActionSet", "actions": actions})

    payload = {
        "type": "AdaptiveCard",
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "version": "1.4",
        "body": body,
    }
    try:
        resp = http_requests.post(TEAMS_WEBHOOK_URL, json=payload, timeout=30)
        if resp.status_code in (200, 202):
            log.info("Teams channel post sent successfully")
        else:
            log.warning(f"Teams webhook post failed ({resp.status_code}): {resp.text[:200]}")
    except Exception as exc:
        log.warning(f"Teams webhook post error: {exc}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    start, end = get_last_week_range()
    log.info(f"=== Weekly Attendance Report  {start} → {end} ===")

    # 1. Download badge log from D3000
    badge_excel = download_badge_excel(start, end)

    # 2. Authenticate to Microsoft Graph
    token = get_graph_token()

    # 3. Azure AD users + managers
    manager_df = fetch_manager_df(token)

    # 4. SharePoint site ID + DataWatch assignees
    site_id         = get_sharepoint_site_id(token)
    datawatch_names = fetch_datawatch_names(token, site_id) if site_id else set()

    # 5. Process attendance
    unique_days, zero_df, total_weekdays = process_attendance(
        badge_excel, start, end, manager_df, datawatch_names
    )

    # 6. Generate Excel report
    filename     = f"Attendance_Report_{start}_{end}.xlsx"
    report_bytes = generate_report_excel(unique_days, zero_df, start, end)

    # 6b. Generate HTML report
    html_filename = f"Attendance_Report_{start}_{end}.html"
    html_bytes    = generate_report_html(unique_days, zero_df, start, end, total_weekdays)

    # 7. Upload to SharePoint (Excel + HTML) — each week gets its own subfolder
    week_folder = f"{start} to {end}"
    file_url = upload_to_sharepoint(token, site_id, filename, report_bytes, week_folder=week_folder) if site_id else ""
    html_url = (
        upload_to_sharepoint(token, site_id, html_filename, html_bytes, content_type="text/html", week_folder=week_folder)
        if site_id else ""
    )

    # 8. Email report (Excel + HTML attached)
    send_email_report(
        token, REPORT_FROM_EMAIL, REPORT_TO_EMAILS,
        unique_days, zero_df, total_weekdays,
        start, end, file_url, filename, report_bytes,
        html_bytes=html_bytes, html_filename=html_filename,
    )

    # 9. Post summary to Teams channel
    post_to_teams_webhook(
        unique_days, zero_df, total_weekdays, start, end, file_url, html_url
    )

    log.info("=== Done ===")


if __name__ == "__main__":
    main()
