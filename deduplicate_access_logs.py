import openpyxl
from datetime import datetime
from collections import defaultdict, Counter

INPUT_FILE = "/Users/yousseffrangieh/Downloads/Export (4).xlsx"
OUTPUT_FILE = "/Users/yousseffrangieh/Downloads/Export_Unique_Days.xlsx"

# Load source
wb_in = openpyxl.load_workbook(INPUT_FILE)
ws_in = wb_in.active

# Collect unique (name, date) pairs
seen = set()
unique_rows = []

for row in ws_in.iter_rows(min_row=2, values_only=True):
    first_name = row[3]  # Col D
    last_name  = row[4]  # Col E
    dt         = row[5]  # Col F - Date Time

    if dt is None:
        continue

    date_only = dt.date() if isinstance(dt, datetime) else dt

    key = (first_name, last_name, date_only)
    if key not in seen:
        seen.add(key)
        unique_rows.append((first_name, last_name, date_only))

# Sort by last name, first name, then date
unique_rows.sort(key=lambda r: (r[1] or "", r[0] or "", r[2]))

# Build output workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Unique Days"

# Header
ws_out.append(["First Name", "Last Name", "Date", "Full Name"])

for first, last, date in unique_rows:
    full = f"{first or ''} {last or ''}".strip()
    ws_out.append([first, last, date, full])

# Format date column
for cell in ws_out["C"][1:]:
    cell.number_format = "YYYY-MM-DD"

# Auto-width columns
for col in ws_out.columns:
    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
    ws_out.column_dimensions[col[0].column_letter].width = max_len + 4

wb_out.save(OUTPUT_FILE)

# Summary stats
days_per_person = Counter((f, l) for f, l, d in unique_rows)
print(f"\nTotal unique person-day entries: {len(unique_rows)}")
print(f"Total unique employees: {len(days_per_person)}")
print(f"\nDays present per employee:")
print(f"{'Name':<40} {'Days':>5}")
print("-" * 47)
for (first, last), days in sorted(days_per_person.items(), key=lambda x: (-x[1], x[0][1] or "")):
    print(f"{(str(first) + ' ' + str(last)):<40} {days:>5}")

print(f"\nOutput saved to: {OUTPUT_FILE}")
